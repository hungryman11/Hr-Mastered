"""Conservative importer for the Infinity staff workbook.

It never creates Zoho identities, guesses email addresses, or silently creates a
company.  Use ``--dry-run`` first; ``--commit`` remains deliberately strict and
only imports rows whose company-owned Department and Position already exist.
"""

from collections import Counter
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify

from core.models import Company, Department, Employee, EmployeeRole, PayrollProfile, Position


REQUIRED_COLUMNS = {'STAFF ID NO.', 'FULL NAME', 'DESIGNATION', 'DEPARTMENT', 'EMAIL', 'STATUS'}
EMPTY_EMAIL_VALUES = {'', 'N/A', 'NA', 'NONE', 'NULL', '-'}


def normalise_email(value):
    value = str(value or '').strip().lower().rstrip(',')
    return value if '@' in value and value.rsplit('@', 1)[1].count('.') >= 1 else ''


def split_name(value):
    parts = [part for part in str(value or '').strip().split() if part]
    return (parts[0], ' '.join(parts[1:])) if parts else ('', '')


class Command(BaseCommand):
    help = 'Reconcile an Infinity staff Excel workbook. Defaults to no database changes.'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', help='Path to the Infinity staff .xlsx workbook')
        parser.add_argument('--company', required=True, help='Existing company name to reconcile against')
        mode = parser.add_mutually_exclusive_group(required=True)
        mode.add_argument('--dry-run', action='store_true', help='Validate and report without writing data')
        mode.add_argument('--commit', action='store_true', help='Apply only fully mapped, valid rows after a reviewed dry run')

    def _load_rows(self, xlsx_path):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError('openpyxl is required to process .xlsx files.') from exc
        path = Path(xlsx_path)
        if path.suffix.lower() != '.xlsx' or not path.is_file():
            raise CommandError('Provide an existing .xlsx workbook path.')
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            headers = [str(cell.value or '').strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            missing = REQUIRED_COLUMNS.difference(headers)
            if missing:
                raise CommandError(f'Workbook is missing required columns: {", ".join(sorted(missing))}')
            return [{headers[index]: cell for index, cell in enumerate(row)} for row in sheet.iter_rows(min_row=2, values_only=True)]
        finally:
            workbook.close()

    def handle(self, *args, **options):
        company = Company.objects.filter(name=options['company']).first()
        if not company:
            raise CommandError('Company must already exist; this importer never creates companies.')
        rows = self._load_rows(options['xlsx_path'])
        report = Counter(total_rows=len(rows))
        valid_rows = []
        seen_emails = set()
        known_departments = {name.lower(): obj for name, obj in Department.objects.filter(company=company).values_list('name', 'id')}
        known_positions = {title.lower(): obj for title, obj in Position.objects.filter(company=company).values_list('title', 'id')}
        for row in rows:
            staff_id = str(row.get('STAFF ID NO.') or '').strip()
            email = normalise_email(row.get('EMAIL'))
            department_name = str(row.get('DEPARTMENT') or '').strip()
            designation = str(row.get('DESIGNATION') or '').strip()
            if not email:
                report['missing_or_invalid_email'] += 1
                continue
            if email in seen_emails:
                report['duplicate_email_in_workbook'] += 1
                continue
            seen_emails.add(email)
            if not staff_id:
                report['missing_staff_id'] += 1
                continue
            if department_name.lower() not in known_departments:
                report['unmapped_department'] += 1
                continue
            if designation.lower() not in known_positions:
                report['unmapped_position'] += 1
                continue
            valid_rows.append((row, email, staff_id, known_departments[department_name.lower()], known_positions[designation.lower()]))
        report['fully_mapped_rows'] = len(valid_rows)
        report['existing_employees_by_email'] = Employee.objects.filter(company=company, email__in=[item[1] for item in valid_rows]).count()
        report['existing_staff_ids'] = PayrollProfile.objects.filter(company=company, employee_number__in=[item[2] for item in valid_rows]).count()
        self.stdout.write(f'Infinity staff reconciliation for: {company.name}')
        for name in ('total_rows', 'fully_mapped_rows', 'missing_or_invalid_email', 'duplicate_email_in_workbook', 'missing_staff_id', 'unmapped_department', 'unmapped_position', 'existing_employees_by_email', 'existing_staff_ids'):
            self.stdout.write(f'{name}={report[name]}')
        self.stdout.write('Mapping: STAFF ID NO. -> PayrollProfile.employee_number; EMAIL -> Employee.email; DESIGNATION -> existing Position; DEPARTMENT -> existing Department.')
        self.stdout.write('Unsupported/no-write fields: branch location, gender, phone number, and Zoho identity.')
        if options['dry_run']:
            self.stdout.write(self.style.WARNING('Dry run complete: no records were created or updated.'))
            return
        if report['missing_or_invalid_email'] or report['duplicate_email_in_workbook'] or report['missing_staff_id'] or report['unmapped_department'] or report['unmapped_position']:
            raise CommandError('Commit aborted: resolve all reconciliation findings before importing.')
        with transaction.atomic():
            created = updated = 0
            for row, email, staff_id, department_id, position_id in valid_rows:
                first_name, last_name = split_name(row.get('FULL NAME'))
                employee = Employee.objects.filter(company=company, email=email).first()
                if employee is None:
                    base_username = slugify(email.split('@', 1)[0]).replace('-', '.')[:140] or 'staff'
                    username = base_username
                    suffix = 2
                    while Employee.objects.filter(username=username).exists():
                        username = f'{base_username[:140]}-{suffix}'
                        suffix += 1
                    employee = Employee(username=username, email=email, company=company, role=EmployeeRole.EMPLOYEE)
                    created += 1
                else:
                    updated += 1
                employee.first_name, employee.last_name = first_name, last_name
                employee.department_id, employee.position_id = department_id, position_id
                employee.org_unit = employee.position.org_unit
                employee.is_active = str(row.get('STATUS') or '').strip().upper() not in {'INACTIVE', 'TERMINATED', 'SUSPENDED'}
                employee.save()
                PayrollProfile.objects.update_or_create(
                    company=company, employee=employee,
                    defaults={'employee_number': staff_id, 'base_salary': 0, 'employment_status': PayrollProfile.EmploymentStatus.ACTIVE if employee.is_active else PayrollProfile.EmploymentStatus.INACTIVE, 'hire_date': employee.date_joined.date()},
                )
        self.stdout.write(self.style.SUCCESS(f'Import complete: created={created}, updated={updated}.'))
